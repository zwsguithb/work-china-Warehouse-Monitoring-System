# -*- coding: utf-8 -*-
"""Excel 导入：领星 ERP 销售/库存导出、生命周期上传。

导入模板字段（首行表头，列名不区分大小写）：
  销售表：sku, platform, d3, d7, d14, d30, d60     （platform ∈ amazon/walmart/other/temu）
  库存表：sku, warehouse, quantity                  （warehouse ∈ 中国仓/temu/amazon_fba/walmart/other）
  生命周期：style_code, lifecycle
"""
import openpyxl
from .database import get_conn, DEFAULT_LIFECYCLE
from datetime import datetime

PLATFORM_ALIAS = {
    "amazon": "amazon", "亚马逊": "amazon",
    "walmart": "walmart", "沃尔玛": "walmart",
    "other": "other", "其他": "other", "其他平台": "other",
    "temu": "temu",
}
WAREHOUSE_ALIAS = {
    "中国仓": "中国仓", "china": "中国仓",
    "temu": "temu", "沃尔玛": "walmart", "walmart": "walmart",
    "amazon": "amazon_fba", "亚马逊": "amazon_fba", "amazon_fba": "amazon_fba",
    "其他": "other", "other": "other",
}


def _header_map(ws):
    """返回 规范化列名 -> 列号(0-based)。"""
    headers = [str(c.value).strip().lower() if c.value else "" for c in ws[1]]
    return {h: i for i, h in enumerate(headers) if h}


def _col(row, hmap, name):
    idx = hmap.get(name.lower())
    if idx is None:
        return None
    v = row[idx].value
    return v


def import_sales(path):
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb.active
    hmap = _header_map(ws)
    conn = get_conn()
    c = conn.cursor()
    n = 0
    for r in range(2, ws.max_row + 1):
        row = ws[r]
        sku = _col(row, hmap, "sku")
        if not sku:
            continue
        platform = PLATFORM_ALIAS.get(str(_col(row, hmap, "platform") or "").strip().lower(), None)
        if not platform:
            continue
        style = str(sku).split("-")[0].rstrip("Bb") if "-" in str(sku) else str(sku)
        c.execute("INSERT OR IGNORE INTO skus(sku, style_code) VALUES(?,?)", (str(sku), style))
        c.execute(
            """INSERT OR REPLACE INTO sales_agg(sku, platform, d3,d7,d14,d30,d60)
               VALUES(?,?,?,?,?,?,?)""",
            (str(sku), platform,
             _num(_col(row, hmap, "d3")), _num(_col(row, hmap, "d7")),
             _num(_col(row, hmap, "d14")), _num(_col(row, hmap, "d30")),
             _num(_col(row, hmap, "d60"))),
        )
        n += 1
    conn.commit()
    conn.close()
    return n


def import_inventory(path):
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb.active
    hmap = _header_map(ws)
    conn = get_conn()
    c = conn.cursor()
    n = 0
    for r in range(2, ws.max_row + 1):
        row = ws[r]
        sku = _col(row, hmap, "sku")
        if not sku:
            continue
        wh = WAREHOUSE_ALIAS.get(str(_col(row, hmap, "warehouse") or "").strip(), None)
        if not wh:
            continue
        style = str(sku).split("-")[0].rstrip("Bb") if "-" in str(sku) else str(sku)
        c.execute("INSERT OR IGNORE INTO skus(sku, style_code) VALUES(?,?)", (str(sku), style))
        c.execute(
            "INSERT OR REPLACE INTO inventory(sku, warehouse, quantity) VALUES(?,?,?)",
            (str(sku), wh, _num(_col(row, hmap, "quantity"))),
        )
        n += 1
    conn.commit()
    conn.close()
    return n


def import_lifecycle(path):
    """上传生命周期；未出现的款号保持/默认淘汰期（不覆盖为淘汰期，仅在不存时默认）。"""
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb.active
    hmap = _header_map(ws)
    conn = get_conn()
    c = conn.cursor()
    now = datetime.now().strftime("%Y-%m-%d %H:%M")
    n = 0
    for r in range(2, ws.max_row + 1):
        row = ws[r]
        sc = _col(row, hmap, "style_code")
        lc = _col(row, hmap, "lifecycle")
        if not sc or not lc:
            continue
        lc = str(lc).strip()
        # 上传即视为明确生命周期，覆盖旧值（is_default=0）
        c.execute(
            """INSERT INTO styles(style_code, lifecycle, is_default, updated_at)
               VALUES(?,?,0,?)
               ON CONFLICT(style_code) DO UPDATE SET lifecycle=excluded.lifecycle,
               is_default=0, updated_at=excluded.updated_at""",
            (str(sc).strip(), lc, now),
        )
        n += 1
    conn.commit()
    conn.close()
    return n


def ensure_default_lifecycle():
    """对尚未在 styles 表中出现的款号，补充一行默认淘汰期(is_default=1)。"""
    conn = get_conn()
    c = conn.cursor()
    styles = c.execute("SELECT DISTINCT style_code FROM skus").fetchall()
    for r in styles:
        sc = r["style_code"]
        exists = c.execute("SELECT 1 FROM styles WHERE style_code=?", (sc,)).fetchone()
        if not exists:
            c.execute("INSERT INTO styles(style_code, lifecycle, is_default) VALUES(?,?,1)",
                      (sc, DEFAULT_LIFECYCLE))
    conn.commit()
    conn.close()


def _num(v):
    if v is None or v == "":
        return 0
    try:
        return float(v)
    except (ValueError, TypeError):
        return 0
