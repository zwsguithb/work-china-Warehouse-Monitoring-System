# -*- coding: utf-8 -*-
"""演示数据：生成示例销售/库存/生命周期，并写出示例 Excel 供上传测试。"""
import random
from datetime import datetime
from pathlib import Path
from .database import get_conn, init_db, DEFAULT_LIFECYCLE
from .data_import import ensure_default_lifecycle

BASE_DIR = Path(__file__).resolve().parent.parent
SAMPLE_DIR = BASE_DIR / "sample_data"

STYLE_POOL = [
    ("2233", "爆旺期"), ("3141", "爆旺期"), ("3142", "热销期"), ("2121", "热销期"),
    ("2122", "平销期"), ("3088", "平销期"), ("2588", "观察期"), ("3180", "观察期"),
    ("4110", "新品期"), ("4088", "新品期"), ("3851", "平销期"), ("3862", "观察期"),
    ("5088", "淘汰期"), ("2021", "淘汰期"), ("3688", "退市期"), ("3086", "已淘汰"),
    ("9110", "暂停"), ("2127", "平销期"), ("2110", "观察期"), ("3841", "新品期"),
    ("6880", "观察期"), ("7719", "淘汰期"), ("8211", "退市期"), ("L605", "淘汰期"),
    ("L613", "已淘汰"),
]
SIZES = ["32A", "32B", "32C", "32D", "34B", "34C", "36B", "36C"]
PLATFORMS = ["amazon", "walmart", "other", "temu"]


def _rand_windows(base):
    """根据基础日销生成 3/7/14/30/60 天窗口的近似聚合量（带噪声）。"""
    d30 = base * 30
    return (
        round(base * 3 * random.uniform(0.8, 1.2), 0),
        round(base * 7 * random.uniform(0.85, 1.15), 0),
        round(base * 14 * random.uniform(0.9, 1.1), 0),
        round(d30 * random.uniform(0.9, 1.1), 0),
        round(base * 60 * random.uniform(0.9, 1.1), 0),
    )


def seed_demo():
    random.seed(20260817)
    init_db()
    conn = get_conn()
    c = conn.cursor()
    # 清空
    for t in ("sales_agg", "inventory", "skus", "styles"):
        c.execute(f"DELETE FROM {t}")
    # 默认配置
    defaults = [
        ("w3", "0.05"), ("w7", "0.15"), ("w14", "0.27"), ("w30", "0.28"), ("w60", "0.25"),
        ("th_baowang", "90"), ("th_rexiao", "60"), ("th_pingxiao", "45"),
        ("th_guancha", "45"), ("th_xinpin", "35"),
        ("top_eliminated_inv", "10"), ("top_eliminated_60", "10"),
        ("top_temu", "5"), ("top_amazon", "20"),
    ]
    c.executemany("INSERT OR REPLACE INTO config(param_key,param_value) VALUES(?,?)", defaults)

    now = datetime.now().strftime("%Y-%m-%d %H:%M")
    # 大多数款号上传生命周期；留 2 个不传以演示默认淘汰期
    uploaded = [s for s in STYLE_POOL if s[0] not in ("L605", "L613")]
    for sc, lc in uploaded:
        c.execute("INSERT OR IGNORE INTO styles(style_code,lifecycle,is_default,updated_at) VALUES(?,?,0,?)",
                  (sc, lc, now))

    for sc, lc in STYLE_POOL:
        n_sku = random.randint(2, 4)
        chosen = random.sample(SIZES, n_sku)
        # 基础日销：爆旺/热销高，淘汰低
        if lc in ("爆旺期", "热销期"):
            base = random.uniform(8, 25)
        elif lc in ("平销期", "观察期", "新品期"):
            base = random.uniform(2, 8)
        else:
            base = random.uniform(0.2, 2)
        for size in chosen:
            sku = f"{sc}B-{size}"
            c.execute("INSERT OR IGNORE INTO skus(sku,style_code) VALUES(?,?)", (sku, sc))
            for p in PLATFORMS:
                w = _rand_windows(base if p != "temu" else base * 0.6)
                c.execute(
                    """INSERT OR REPLACE INTO sales_agg(sku,platform,d3,d7,d14,d30,d60)
                       VALUES(?,?,?,?,?,?,?)""", (sku, p, *w))
            # 库存：中国仓 + temu
            china = round(random.uniform(20, 600) if lc not in ("淘汰期", "已淘汰", "退市期", "暂停") else random.uniform(0, 80))
            temu = round(random.uniform(0, 200))
            c.execute("INSERT OR REPLACE INTO inventory(sku,warehouse,quantity) VALUES(?,?,?)",
                      (sku, "中国仓", china))
            c.execute("INSERT OR REPLACE INTO inventory(sku,warehouse,quantity) VALUES(?,?,?)",
                      (sku, "temu", temu))
    conn.commit()
    conn.close()
    ensure_default_lifecycle()
    print("Demo data seeded.")


def make_sample_excels():
    """生成示例上传 Excel（销售/库存/生命周期），位于 sample_data/。"""
    import openpyxl

    SAMPLE_DIR.mkdir(exist_ok=True)
    conn = get_conn()
    c = conn.cursor()

    # 销售表
    ws = openpyxl.Workbook().active
    ws.append(["sku", "platform", "d3", "d7", "d14", "d30", "d60"])
    for r in c.execute("SELECT sku,platform,d3,d7,d14,d30,d60 FROM sales_agg"):
        ws.append([r["sku"], r["platform"], r["d3"], r["d7"], r["d14"], r["d30"], r["d60"]])
    ws.parent.save(SAMPLE_DIR / "lingxing_sales_sample.xlsx")

    # 库存表
    ws2 = openpyxl.Workbook().active
    ws2.append(["sku", "warehouse", "quantity"])
    for r in c.execute("SELECT sku,warehouse,quantity FROM inventory"):
        ws2.append([r["sku"], r["warehouse"], r["quantity"]])
    ws2.parent.save(SAMPLE_DIR / "inventory_sample.xlsx")

    # 生命周期表（取已上传的部分）
    ws3 = openpyxl.Workbook().active
    ws3.append(["style_code", "lifecycle"])
    for r in c.execute("SELECT style_code,lifecycle FROM styles WHERE is_default=0"):
        ws3.append([r["style_code"], r["lifecycle"]])
    ws3.parent.save(SAMPLE_DIR / "lifecycle_sample.xlsx")
    conn.close()
    print("Sample excels written to sample_data/.")


if __name__ == "__main__":
    seed_demo()
    make_sample_excels()
